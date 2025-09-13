import sys
import os
import json
from typing import Dict, List
from datetime import datetime
import re
from openpyxl import load_workbook

from .playwright_ucr import fetch_ucr_fee_sync, parse_ucr_html


def process_json_input(json_data: dict, acctkey: str) -> dict:
    """Process JSON input and return JSON results."""
    results = []
    
    line_items = json_data.get("line_items", [])
    if not line_items:
        return {"error": "No line_items found in input JSON"}
    
    for index, item in enumerate(line_items, start=1):
        # Extract and normalize data from JSON
        zip_code = str(item.get("ZipCode", "")).strip()
        cpt_code = str(item.get("CPTcode", "")).strip()
        date_str = str(item.get("date", "")).strip()
        
        # Normalize date format (convert from YYYY-MM-DD to MM/DD/YYYY)
        def normalize_date(date_input: str) -> str:
            if not date_input:
                return ""
            try:
                # Try YYYY-MM-DD format first
                if "-" in date_input and len(date_input) == 10:
                    dt = datetime.strptime(date_input, "%Y-%m-%d")
                    return dt.strftime("%m/%d/%Y")
                # Try MM/DD/YYYY format
                elif "/" in date_input:
                    dt = datetime.strptime(date_input, "%m/%d/%Y")
                    return dt.strftime("%m/%d/%Y")
                else:
                    return date_input
            except Exception:
                return date_input
        
        service_date = normalize_date(date_str)
        procedure_code = re.sub(r"\D", "", cpt_code).strip()  # digits only
        zip_code = re.sub(r"\D", "", zip_code).strip().zfill(5)  # exactly 5 digits
        
        # Basic validations
        if len(service_date) != 10 or service_date.count('/') != 2:
            error_result = {
                "procedureCode": procedure_code,
                "percentiles": {},
                "currency": "USD",
                "zip_code": int(zip_code) if zip_code.isdigit() else 0,
                "date": service_date,
                "line_number": index,
                "error": f"Invalid date format: '{date_str}' (expected YYYY-MM-DD or MM/DD/YYYY)"
            }
            results.append(error_result)
            print(f"Line {index}: error invalid date '{date_str}'", flush=True)
            continue
            
        if not procedure_code:
            error_result = {
                "procedureCode": "",
                "percentiles": {},
                "currency": "USD",
                "zip_code": int(zip_code) if zip_code.isdigit() else 0,
                "date": service_date,
                "line_number": index,
                "error": "Missing CPT code"
            }
            results.append(error_result)
            print(f"Line {index}: error missing CPT", flush=True)
            continue
            
        if not re.match(r"^\d{5}$", zip_code):
            error_result = {
                "procedureCode": procedure_code,
                "percentiles": {},
                "currency": "USD",
                "zip_code": int(zip_code) if zip_code.isdigit() else 0,
                "date": service_date,
                "line_number": index,
                "error": f"Invalid ZIP code: '{zip_code}' (expected 5 digits)"
            }
            results.append(error_result)
            print(f"Line {index}: error invalid ZIP '{zip_code}'", flush=True)
            continue
        
        print(f"Line {index}: date={service_date} cpt={procedure_code} zip={zip_code} ...", flush=True)
        
        # Scrape UCR data
        html, err = fetch_ucr_fee_sync(
            acct_key=acctkey,
            service_date=service_date,
            procedure_code=procedure_code,
            zip_code=zip_code,
            percentile="50",
            timeout_ms=20000,
        )
        
        if not err and html:
            parsed = parse_ucr_html(html)
            
            # Create result without description field
            result = {
                "procedureCode": parsed.get("procedureCode", procedure_code),
                "percentiles": parsed.get("percentiles", {}),
                "currency": parsed.get("currency", "USD"),
                "zip_code": int(zip_code),
                "date": service_date,
                "line_number": index
            }
            
            # Add error field if no percentiles found
            if not parsed.get("percentiles"):
                result["error"] = "No percentiles found in response"
            
            results.append(result)
            print(f"Line {index}: {json.dumps(result, indent=2)}", flush=True)
            
        else:
            # Create error result
            error_result = {
                "procedureCode": procedure_code,
                "percentiles": {},
                "currency": "USD",
                "zip_code": int(zip_code),
                "date": service_date,
                "line_number": index,
                "error": f"Scraping failed: {err}" if err else "Unknown error"
            }
            results.append(error_result)
            print(f"Line {index}: error {err}", flush=True)
    
    return {
        "results": results,
        "total_processed": len(results),
        "successful": len([r for r in results if not r.get("error")]),
        "failed": len([r for r in results if r.get("error")])
    }


def process_to_json(input_path: str, acctkey: str) -> str:
    """Process Excel input and output JSON results instead of Excel."""
    wb = load_workbook(input_path)
    ws = wb.active
    
    results: List[Dict] = []
    
    # Iterate rows starting from 2 (assuming row 1 is header)
    row = 2
    while True:
        date_v = ws[f"A{row}"].value
        cpt_v = ws[f"B{row}"].value
        zip_v = ws[f"C{row}"].value
        if not date_v and not cpt_v and not zip_v:
            break

        if date_v and cpt_v and zip_v:
            # Normalize date to strict MM/DD/YYYY (10 chars with slashes)
            def coerce_date(val) -> str:
                if isinstance(val, datetime):
                    return val.strftime("%m/%d/%Y")
                s = str(val).strip()
                for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d"):
                    try:
                        return datetime.strptime(s, fmt).strftime("%m/%d/%Y")
                    except Exception:
                        continue
                # last resort: try M/D/YYYY and pad
                m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
                if m:
                    mm = m.group(1).zfill(2)
                    dd = m.group(2).zfill(2)
                    yy = m.group(3)
                    return f"{mm}/{dd}/{yy}"
                return s

            service_date = coerce_date(date_v)
            procedure_code = re.sub(r"\D", "", str(cpt_v)).strip()  # digits only
            zip_code = re.sub(r"\D", "", str(zip_v)).strip().zfill(5)  # exactly 5 digits

            # Basic validations per site requirements
            if len(service_date) != 10 or service_date.count('/') != 2:
                print(f"Row {row}: error invalid date '{service_date}'", flush=True)
                row += 1
                continue
            if not procedure_code:
                print(f"Row {row}: error missing CPT", flush=True)
                row += 1
                continue
            if not re.match(r"^\d{5}$", zip_code):
                print(f"Row {row}: error invalid ZIP '{zip_code}'", flush=True)
                row += 1
                continue

            print(f"Row {row}: date={service_date} cpt={procedure_code} zip={zip_code} ...", flush=True)
            html, err = fetch_ucr_fee_sync(
                acct_key=acctkey,
                service_date=service_date,
                procedure_code=procedure_code,
                zip_code=zip_code,
                percentile="50",
                timeout_ms=20000,
            )
            
            if not err and html:
                parsed = parse_ucr_html(html)
                
                # Create enhanced JSON result with additional fields
                result = {
                    "procedureCode": parsed.get("procedureCode", procedure_code),
                    "percentiles": parsed.get("percentiles", {}),
                    "currency": parsed.get("currency", "USD"),
                    "zip_code": int(zip_code),
                    "date": service_date,
                    "row_number": row
                }
                
                # Add error field if no percentiles found
                if not parsed.get("percentiles"):
                    result["error"] = "No percentiles found in response"
                
                results.append(result)
                print(f"Row {row}: {json.dumps(result, indent=2)}", flush=True)
                
                # Save HTML snapshot for debugging
                try:
                    snap_dir = os.path.dirname(input_path)
                    snap_name = f"row_{row}_{procedure_code}_{zip_code}.html"
                    with open(os.path.join(snap_dir, snap_name), "w", encoding="utf-8") as f:
                        f.write(html)
                except Exception:
                    pass
            else:
                # Create error result
                error_result = {
                    "procedureCode": procedure_code,
                    "percentiles": {},
                    "currency": "USD",
                    "zip_code": int(zip_code),
                    "date": service_date,
                    "row_number": row,
                    "error": f"Scraping failed: {err}" if err else "Unknown error"
                }
                results.append(error_result)
                print(f"Row {row}: error {err}", flush=True)

        row += 1

    # Save results to JSON file
    output_path = os.path.splitext(input_path)[0] + "_results.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    
    print(f"Wrote JSON results to: {output_path}", flush=True)
    print(f"Total rows processed: {len(results)}", flush=True)
    return output_path


def fill_sheet(input_path: str, acctkey: str) -> str:
    wb = load_workbook(input_path)
    ws = wb.active

    # Ensure headers for percentile columns exist
    headers: Dict[str, str] = {
        "50": "D",
        "55": "E",
        "60": "F",
        "65": "G",
        "70": "H",
        "75": "I",
        "80": "J",
        "85": "K",
        "90": "L",
        "95": "M",
    }
    for pct, col in headers.items():
        if not ws[f"{col}1"].value:
            ws[f"{col}1"] = pct

    # Iterate rows starting from 2 (assuming row 1 is header)
    row = 2
    while True:
        date_v = ws[f"A{row}"].value
        cpt_v = ws[f"B{row}"].value
        zip_v = ws[f"C{row}"].value
        if not date_v and not cpt_v and not zip_v:
            break

        if date_v and cpt_v and zip_v:
            # Normalize date to strict MM/DD/YYYY (10 chars with slashes)
            def coerce_date(val) -> str:
                if isinstance(val, datetime):
                    return val.strftime("%m/%d/%Y")
                s = str(val).strip()
                for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d"):
                    try:
                        return datetime.strptime(s, fmt).strftime("%m/%d/%Y")
                    except Exception:
                        continue
                # last resort: try M/D/YYYY and pad
                m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
                if m:
                    mm = m.group(1).zfill(2)
                    dd = m.group(2).zfill(2)
                    yy = m.group(3)
                    return f"{mm}/{dd}/{yy}"
                return s

            service_date = coerce_date(date_v)
            procedure_code = re.sub(r"\D", "", str(cpt_v)).strip()  # digits only
            zip_code = re.sub(r"\D", "", str(zip_v)).strip().zfill(5)  # exactly 5 digits

            # Basic validations per site requirements
            if len(service_date) != 10 or service_date.count('/') != 2:
                print(f"Row {row}: error invalid date '{service_date}'", flush=True)
                row += 1
                continue
            if not procedure_code:
                print(f"Row {row}: error missing CPT", flush=True)
                row += 1
                continue
            if not re.match(r"^\d{5}$", zip_code):
                print(f"Row {row}: error invalid ZIP '{zip_code}'", flush=True)
                row += 1
                continue

            print(f"Row {row}: date={service_date} cpt={procedure_code} zip={zip_code} ...", flush=True)
            html, err = fetch_ucr_fee_sync(
                acct_key=acctkey,
                service_date=service_date,
                procedure_code=procedure_code,
                zip_code=zip_code,
                percentile="50",
                timeout_ms=20000,
            )
            if not err and html:
                parsed = parse_ucr_html(html)
                # Save snapshot for auditing
                try:
                    snap_dir = os.path.dirname(input_path)
                    snap_name = f"row_{row}_{procedure_code}_{zip_code}.html"
                    with open(os.path.join(snap_dir, snap_name), "w", encoding="utf-8") as f:
                        f.write(html)
                except Exception:
                    pass
                per = parsed.get("percentiles", {}) if isinstance(parsed.get("percentiles"), dict) else {}
                for pct, col in headers.items():
                    if pct in per:
                        ws[f"{col}{row}"] = per[pct]
                print(f"Row {row}: percentiles={json.dumps(per)}", flush=True)
            else:
                ws[f"{headers['50']}{row}"] = f"ERR: {err}" if err else "ERR"
                print(f"Row {row}: error {err}", flush=True)

        row += 1

    out_path = os.path.splitext(input_path)[0] + "_filled.xlsx"
    wb.save(out_path)
    print(f"Wrote: {out_path}", flush=True)
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python -m chatbot.services.ucr_batch_runner <input_path> <acctkey> [--json] [--json-input]")
        print("  --json: Output JSON format instead of Excel")
        print("  --json-input: Input is JSON file instead of Excel")
        sys.exit(1)
    
    input_path = sys.argv[1]
    acct = sys.argv[2]
    output_json = "--json" in sys.argv
    json_input = "--json-input" in sys.argv
    
    if json_input:
        # Process JSON input
        try:
            with open(input_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            result = process_json_input(json_data, acct)
            print(json.dumps(result, indent=2))
        except Exception as e:
            print(f"Error processing JSON input: {e}")
            sys.exit(1)
    elif output_json:
        # Process Excel input, output JSON
        result = process_to_json(input_path, acct)
        print(f"JSON output: {result}")
    else:
        # Process Excel input, output Excel
        result = fill_sheet(input_path, acct)
        print(f"Excel output: {result}")


    